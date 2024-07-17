"""
Takes in two line arguments:
1 - original csv file to be compared against
2 - seconday csv file which should match the first ASIDE from order
Compares both for discrepencies. If any are found they are noted into a file 'issues.txt' with their location in Excel format. 

Assumptions:
- All fields should be the same
- The unique identifier of each row is the first column 
"""
import csv 
import sys
import os.path
from tqdm import tqdm
import time
from enum import Enum
from openpyxl import Workbook, styles

class NATURE_OF_ISSUES(Enum): 
    OK = "No issues found."
    DISCREPENCY = "Discrepencies found, please check the issue log."
    FIELDS_MISMATCH = "Columns fields don't match, please check the issue log."
    FIELDS_LENGTH_MISMATCH = "The number of columns in each csv file don't match. Please check for any trailing commas with notepad."
    
class ISSUE_ITEM:
    def __init__(self, original_row: list, uploaded_row: list, mismatched_columns_indexes: list) -> None:
        self.original_row = original_row
        self.uploaded_row = uploaded_row
        self.mismatched_columns_indexes = mismatched_columns_indexes
    
class ISSUES_MAIN:
    def __init__(self) -> None:
        self.status = NATURE_OF_ISSUES.OK
        self.issue_list: list[ISSUES_MAIN] = []
        self.uploaded_fields: list[str] = []
        self.original_fields: list[str] = []
        
    def insert_issue_missing_uploaded_row(self, original_row: list, value_of_identifier: str, column_of_identifier: int = 0):
        row_to_indicate_missing_row = []
        for i in range(len(original_row)):
            if i != column_of_identifier:
                row_to_indicate_missing_row.append('MISSING')
            else: 
                row_to_indicate_missing_row.append(value_of_identifier)
        issue_item = ISSUE_ITEM(original_row, row_to_indicate_missing_row, [column_of_identifier])
        self.issue_list.append(issue_item)
        return issue_item

    def insert_issue(self, original_row: list, uploaded_row: list, columns_where_discrepency_is_found: list):
        issue_item = ISSUE_ITEM(original_row, uploaded_row, columns_where_discrepency_is_found)
        self.issue_list.append(issue_item)
        return issue_item

def get_dir(dir: str):
    if dir == "":
        return dir
    if not os.path.isdir(dir): 
        os.mkdir(dir)
    return f'{dir}/'

def find_discrepencies(uploaded_file_path: str, 
                       original_file_path: str, 
                       progress_to_show_in_gui = None,
                       status_to_show_in_gui = None,
                       uploaded_file_identifiying_field_index: int = 0,
                       original_file_identifiying_field_index: int = 0) -> ISSUES_MAIN:
    """
    Finds the difference between two CSV files. 

    Args: 
        uploaded_file_path: The path that points to the CSV file which should match the original. 
        original_file_path: The path that points to the CSV file which is meant to be compared against (or in other words, this CSV file is the source of truth).
        progress_to_show_in_gui: A function from the GUI Python file which accepts the current percentage progress of the script as an integer.

    Returns: 
        A list containing strings of the discrepencies found.
    """
    # Update status
    if status_to_show_in_gui:
        status_to_show_in_gui('Setting up...')

    # Read uploaded csv file
    f_uploaded = open(uploaded_file_path, 'r')
    uploaded_csv_reader = csv.reader(f_uploaded)

    # Field names 
    fields_uploaded_csv = next(uploaded_csv_reader)

    # Read original csv file
    f_ori = open(original_file_path, 'r')
    ori_csv_reader = csv.reader(f_ori)
    
    # Field names 
    fields_ori_csv = next(ori_csv_reader)

    # Initialise issues custom class to hold all the found issues (if any)
    issues = ISSUES_MAIN()

    # Save the fields for later use 
    issues.original_fields = fields_ori_csv
    issues.uploaded_fields = fields_uploaded_csv

    # Checking if every field in the original csv exists in the uploaded csv - If failed will not proceed with rest of check, return issue log immediately
    mismatched_fields = []
    for i in range(len(fields_ori_csv)):
        if fields_ori_csv[i] not in fields_uploaded_csv:
            issues.status = NATURE_OF_ISSUES.FIELDS_MISMATCH
            mismatched_fields.append(i)
    if issues.status == NATURE_OF_ISSUES.FIELDS_MISMATCH: 
        issues.insert_issue(fields_ori_csv, fields_uploaded_csv, mismatched_fields)
        return issues
    if len(fields_ori_csv) != len(fields_uploaded_csv): 
        issues.status = NATURE_OF_ISSUES.FIELDS_LENGTH_MISMATCH
        return issues
    
    # Update status
    if status_to_show_in_gui:
        status_to_show_in_gui('Caching uploaded csv...')

    # Hash the uploaded csv file based on its key value
    uploaded_hashed_csv = {}
    for row in uploaded_csv_reader:
        uploaded_hashed_csv[row[uploaded_file_identifiying_field_index]] = row

    # Close the uploaded csv file, it's no longer needed as its now been hashed into memory
    f_uploaded.close()

    # Update status
    if status_to_show_in_gui:
        status_to_show_in_gui('Comparing csvs...')

    # For each row in the original csv file
    # Check if it exists in the uploaded csv file
    # If it exists, compare that row of values to the original 
    previous_update = 0 # For GUI 
    pbar = tqdm(total=len(uploaded_hashed_csv), desc="Comparing rows")
    for row_num, row_from_ori_csv in enumerate( ori_csv_reader ):
        if row_from_ori_csv[original_file_identifiying_field_index] not in uploaded_hashed_csv:
            issues.insert_issue_missing_uploaded_row(row_from_ori_csv, row_from_ori_csv[original_file_identifiying_field_index], original_file_identifiying_field_index)
            continue

        row_from_uploaded_csv = uploaded_hashed_csv[row_from_ori_csv[original_file_identifiying_field_index]]

        mismatched_fields = []
        for col_num in range(len(row_from_ori_csv)):
            if row_from_uploaded_csv[col_num] != row_from_ori_csv[col_num]:
                mismatched_fields.append(col_num)
        if len(mismatched_fields) > 0: 
            issues.insert_issue(row_from_ori_csv,row_from_uploaded_csv,mismatched_fields)
        
        # GUI and TQDM progress bars
        pbar.update(1)
        progress_in_percentage = min( int(row_num / len(uploaded_hashed_csv)*100) + 1, 100)
        if progress_to_show_in_gui != None and previous_update != progress_in_percentage:
            progress_to_show_in_gui(progress_in_percentage)
            previous_update = progress_in_percentage

    # Close the original csv file, we've read through everything 
    f_ori.close()

    # Write the issues into an excel file and return the issue log
    if len(issues.issue_list) == 0: 
        issues.status = NATURE_OF_ISSUES.OK
    else: 
        issues.status = NATURE_OF_ISSUES.DISCREPENCY

    # Update status
    if status_to_show_in_gui:
        status_to_show_in_gui(issues.status.value)
    return issues

def write_issues(issues: list[ISSUE_ITEM], output_dir: str = ""):
    if len(issues) == 0:
        return

    output_dir = get_dir(output_dir)

    f = open(f'{output_dir}issues_{time.strftime("%Y_%m_%d_%H_%M_%S", time.gmtime())}.txt', 'a+')
    for row in issues:
        original_row = 'ORI |'
        for x, item in enumerate(row.original_row): 
            if x in row.mismatched_columns_indexes:
                original_row += f' <|{item}|>'
            else:
                original_row += f' {item}'
        uploaded_row = 'UPL |'
        for x, item in enumerate(row.uploaded_row): 
            if x in row.mismatched_columns_indexes:
                uploaded_row += f' <|{item}|>'
            else:
                uploaded_row += f' {item}'
        f.write(f'{original_row}\n')
        f.write(f'{uploaded_row}\n')
        f.write(f'\n')
    f.close()
    return 

def write_issues_to_excel(issues: list[ISSUE_ITEM], fields: list[str], progress_bar = None, output_dir: str = "") -> None:
    if len(issues) == 0:
        return 

    output_dir = get_dir(output_dir)

    # Create the excel file
    filename = f'{output_dir}issues_{time.strftime("%Y_%m_%d_%H_%M_%S", time.gmtime())}.xlsx'
    if not os.path.isfile(filename):
        wb = Workbook()
        wb.save(filename)
    sheet = wb.active # Create a pointer to the current sheet

    for i, item in enumerate(fields):
        item = item.strip().replace("ï»¿", "")
        sheet.cell(row=1, column=i+2).value = item # offset two as the first row is taken up by the original/upload identifier

    row_index = 2
    pbar = tqdm(total=len(issues), desc="Inserting into Excel")
    for i, row in enumerate(issues):
        # Insert the original row, offsetting 1 column to indicate in a new column for the type of row it is
        sheet.cell(row=row_index, column=1).value = 'ORI'
        for x, item in enumerate(row.original_row):
            sheet.cell(row=row_index, column=x+2).value = item
        # Highlight the columns in which the discrepency is found
        for mismatched_column_index in row.mismatched_columns_indexes: 
            sheet.cell(row=row_index, column=mismatched_column_index+2).fill = styles.PatternFill(fill_type="solid", fgColor="AFEEEE")
        row_index += 1 # Move to the next row for the uploaded row

        sheet.cell(row=row_index, column=1).value = 'UPL'
        for x, item in enumerate(row.uploaded_row):
            sheet.cell(row=row_index, column=x+2).value = item
        for x, mismatched_column_index in enumerate(row.mismatched_columns_indexes): 
            sheet.cell(row=row_index, column=mismatched_column_index+2).fill = styles.PatternFill(fill_type="solid", fgColor="AFEEEE")
        row_index += 1

        # Update progress 
        pbar.update(1)
        if progress_bar:
            progress_bar( min( int(i / len(issues)*100) + 1, 100) )

    wb.save(filename)
    return 

def load_mapping_uploaded_to_original():
    f = open('mapping.csv', 'r')
    mapping_csv_reader = csv.reader(f)
    fields = next(mapping_csv_reader)
    mapping = {}
    for _, row in enumerate(mapping_csv_reader):
        original, uploaded = row[0], row[1]
        mapping[original] = uploaded
        mapping[uploaded] = original
    return mapping

if __name__ == "__main__": 
    # Get input parameters, if not use default names and column index for identifier
    try:
        name_of_script = sys.argv[0]
        ori_file_name = sys.argv[1]
        uploaded_file_name = sys.argv[2]
    except: 
        print(f'Using default filenames ori.csv and uploaded.csv with column index 0 as identifier\n\n')
        ori_file_name = 'ori.csv'
        uploaded_file_name = 'uploaded.csv'
    try: 
        uploaded_file_identifiying_field_index = int(sys.argv[3])
        original_file_identifiying_field_index = int(sys.argv[4])
    except:
        uploaded_file_identifiying_field_index = 0
        original_file_identifiying_field_index = 0

    # Give error and quit script if files cannot be found
    if not (os.path.isfile(ori_file_name) and os.path.isfile(uploaded_file_name)):
        raise Exception(f'Files {ori_file_name} or {uploaded_file_name} cannot be found. Terminating script.')
    
    issues = find_discrepencies(uploaded_file_name,
                                 ori_file_name, 
                                 uploaded_file_identifiying_field_index=uploaded_file_identifiying_field_index,
                                 original_file_identifiying_field_index=original_file_identifiying_field_index)
    write_issues(issues.issue_list)
    print(issues.status.value)
