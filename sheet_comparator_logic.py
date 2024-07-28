"""
Takes in two line arguments:
1 - original csv file to be compared against
2 - secondary csv file which should match the first ASIDE from order
Compares both for discrepancies. If any are found they are noted into a file 'issues.txt' with their location in Excel format. 

Assumptions:
- All fields should be the same
- The unique identifier of each row is the first column 
"""
import csv 
import sys
import os.path
import time
from enum import Enum
from openpyxl import Workbook, styles, load_workbook
import threading
from queue import Queue, Empty

class Progress:
    def __init__(self, total_work, progress_bar):
        self.total_work = total_work
        self.completed_work = 0
        self.lock = threading.Lock()
        self.progress_bar = progress_bar
    def update_progress(self):      
        with self.lock:
            if self.progress_bar:
                self.completed_work += 1 
                self.progress_bar(self.completed_work/self.total_work*100)
        return
class NATURE_OF_ISSUES(Enum): 
    OK = "No issues found."
    DISCREPANCY = "Discrepancies found, please check the issue log."
    FIELDS_MISMATCH = "Columns fields don't match, please check the issue log."
    FIELDS_LENGTH_MISMATCH = "The number of columns in each csv file don't match. Please check for any trailing commas with notepad."
    
class ISSUE_ITEM:
    def __init__(self, original_row: list, uploaded_row: list, mismatched_columns_indexes: list) -> None:
        self.original_row = original_row
        self.uploaded_row = uploaded_row
        self.mismatched_columns_indexes = mismatched_columns_indexes
    
class ISSUES_MAIN:
    def __init__(self) -> None:
        self.nature_of_issues: list[NATURE_OF_ISSUES] = [] 
        self.name = ""
        self.issue_list: list[ISSUE_ITEM] = []
        self.original_fields: list[str] = []
        self.uploaded_fields: list[str] =[] 
        self.uploaded_hashed_fields_idxs: dict = {} #original field: uploaded field index 

    def has_issues(self) -> bool:
        return len(self.nature_of_issues) > 0 
    
    def update_uploaded_hashed_fields_idxs(self) -> dict:
        for ori_field in self.original_fields:
            for upl_field_idx, upl_field in enumerate(self.uploaded_fields):
                if upl_field == ori_field:
                    self.uploaded_hashed_fields_idxs[ori_field] = upl_field_idx
        return self.uploaded_hashed_fields_idxs
    
    def insert_issue_missing_uploaded_row(self, original_row: list, value_of_identifier: str, column_of_identifier: int = 0):
        row_to_indicate_missing_row = []
        for i in range(len(original_row)):
            if i != column_of_identifier:
                row_to_indicate_missing_row.append('MISSING')
            else: 
                row_to_indicate_missing_row.append(value_of_identifier)
        issue_item = ISSUE_ITEM(original_row, row_to_indicate_missing_row, [column_of_identifier])
        self.issue_list.append(issue_item)
        return issue_item

    def insert_issue(self, original_row: list, uploaded_row: list, columns_where_discrepancy_is_found: list[int]):
        mapped_uploaded_row = [] 
        for ori_field in self.original_fields:
            upl_field_idx = self.uploaded_hashed_fields_idxs[ori_field]
            value_from_upl_sheet = uploaded_row[upl_field_idx]
            mapped_uploaded_row.append(value_from_upl_sheet)
        issue_item = ISSUE_ITEM(original_row, mapped_uploaded_row, columns_where_discrepancy_is_found)
        self.issue_list.append(issue_item)
        return issue_item

    def insert_missing_column(self, original_columns, uploaded_columns, columns_missing_from_uploaded: list[str]):
        issue_item = ISSUE_ITEM(original_row=original_columns, 
                                 uploaded_row=uploaded_columns, 
                                 mismatched_columns_indexes=columns_missing_from_uploaded)
        self.issue_list.append(issue_item)
        return issue_item

def xlsx_to_csv(excel_file_path: str) -> tuple:
    wb = load_workbook(excel_file_path)
    sh = wb.active.values
    fields = [] # headers of the sheet
    rest = [] # the data/content of the sheet
    for i, row in enumerate(sh): 
        if i == 0: # if its the first row, it is the fields
            fields = [str(cell) for cell in row if cell != None]
        else: # rest is normal data rows
            rest.append([(u"" if cell == None else str(cell)) for cell in row]) # Ensure empty cells are blanks rather than "None" objects, also convert to string to prevent typing discrepancies
    wb.close()
    return fields, rest 

def find_discrepancies(uploaded_file_path: str, 
                       original_file_path: str, 
                       progress_to_show_in_gui = None,
                       status_to_show_in_gui = None,
                       uploaded_file_identifying_field_index: int = 0,
                       original_file_identifying_field_index: int = 0,
                       ignore_leading_and_trailing_whitespaces: bool = False) -> ISSUES_MAIN:
    """
    Finds the difference between two CSV files. 

    Args: 
        uploaded_file_path: The path that points to the CSV file which should match the original. 
        original_file_path: The path that points to the CSV file which is meant to be compared against (or in other words, this CSV file is the source of truth).
        progress_to_show_in_gui: A function from the GUI Python file which accepts the current percentage progress of the script as an integer.

    Returns: 
        A list containing strings of the discrepancies found.
    """
    # Update status
    previous_update = 0 # For GUI 

    # If Excel file given, convert to CSV first
    if uploaded_file_path.split('.')[-1] == 'csv': 
        # Read uploaded csv file
        f_uploaded = open(uploaded_file_path, 'r')
        uploaded_csv_reader = csv.reader(f_uploaded)
        # Field names 
        fields_uploaded_csv = next(uploaded_csv_reader)
    else:
        fields_uploaded_csv, uploaded_csv_reader = xlsx_to_csv(uploaded_file_path)
    if original_file_path.split('.')[-1] == 'csv': 
        # Read original csv file
        f_ori = open(original_file_path, 'r')
        ori_csv_reader = csv.reader(f_ori)
        # Field names 
        fields_ori_csv = next(ori_csv_reader)
    else:
        fields_ori_csv, ori_csv_reader = xlsx_to_csv(original_file_path)

    # Initialise issues custom class to hold all the found issues (if any)
    issues = ISSUES_MAIN()
    issues.name = original_file_path.split('/')[-1].split('.')[0]
    # Save the fields for later use 
    issues.original_fields = fields_ori_csv
    issues.uploaded_fields = fields_uploaded_csv

    # Checking if every field in the original csv exists in the uploaded csv - If failed will not proceed with rest of check, return issue log immediately
    mismatched_fields = []
    for i in range(len(fields_ori_csv)):
        if fields_ori_csv[i] not in fields_uploaded_csv:
            issues.nature_of_issues.append(NATURE_OF_ISSUES.FIELDS_MISMATCH)
            mismatched_fields.append(i)
    if NATURE_OF_ISSUES.FIELDS_MISMATCH in issues.nature_of_issues: 
        issues.insert_missing_column(fields_ori_csv, fields_uploaded_csv, mismatched_fields)
        # TODO Toggle for checking regardless of fields being mismatched
        return issues
    
    # Update status
    if status_to_show_in_gui:
        status_to_show_in_gui(f'Caching uploaded sheet {issues.name}')
    if progress_to_show_in_gui != None:
        progress_to_show_in_gui(0)

    # Hash the uploaded csv file based on its key value
    uploaded_hashed_csv = {}
    for row in uploaded_csv_reader:
        uploaded_hashed_csv[row[uploaded_file_identifying_field_index]] = row

    # Update status
    if progress_to_show_in_gui != None:
        progress_to_show_in_gui(25)

    # Hash the indexes of the uploaded csv fields
    uploaded_hashed_fields_index = issues.update_uploaded_hashed_fields_idxs()

    # Close the uploaded csv file, it's no longer needed as its now been hashed into memory
    try:
        f_uploaded.close()
    except UnboundLocalError: 
        pass

    # Update status
    if status_to_show_in_gui:
        status_to_show_in_gui(f'Comparing sheet {issues.name}')

    # For each row in the original csv file
    # Check if it exists in the uploaded csv file
    # If it exists, compare that row of values to the original 
    # Each ROW 
    for row_num, row_from_ori_csv in enumerate( ori_csv_reader ):
        # Row from original sheet cannot be found in the uploaded sheet
        if row_from_ori_csv[original_file_identifying_field_index] not in uploaded_hashed_csv:
            issues.insert_issue_missing_uploaded_row(row_from_ori_csv, row_from_ori_csv[original_file_identifying_field_index], original_file_identifying_field_index)
            continue

        # Corresponding row from the uploaded sheet 
        row_from_uploaded_csv = uploaded_hashed_csv[row_from_ori_csv[original_file_identifying_field_index]]

        # Each COLUMN (CELL)
        mismatched_fields = []
        for col_num in range(len(fields_ori_csv)):
            cell_from_ori_csv = row_from_ori_csv[col_num]
            cell_from_upl_csv = row_from_uploaded_csv[uploaded_hashed_fields_index[fields_ori_csv[col_num]]]
            if ignore_leading_and_trailing_whitespaces:
                cell_from_ori_csv = cell_from_ori_csv.strip() 
                cell_from_upl_csv = cell_from_upl_csv.strip()
            if cell_from_ori_csv != cell_from_upl_csv:
                mismatched_fields.append(col_num)
        if len(mismatched_fields) > 0: 
            issues.insert_issue(row_from_ori_csv,row_from_uploaded_csv,mismatched_fields)
        
        # GUI progress bars
        progress_in_percentage = min( int(row_num / len(uploaded_hashed_csv)*75) + 25, 100)
        if progress_to_show_in_gui != None and previous_update != progress_in_percentage:
            progress_to_show_in_gui(progress_in_percentage)
            previous_update = progress_in_percentage
    
    # Mark issues found 
    if len(issues.issue_list): issues.nature_of_issues.append(NATURE_OF_ISSUES.DISCREPANCY)  

    # Close the original csv file, we've read through everything 
    try:
        f_ori.close()
    except UnboundLocalError: 
        pass

    # Update status
    if status_to_show_in_gui:
        status_to_show_in_gui(f'Finished comparing {issues.name}')

    return issues

def write_issues(issues: ISSUES_MAIN, output_dir: str = "", use_excel: bool = False, progress_bar = None, progress_status = None) -> None: 
    # Writing logic for a text file
    def write_to_text():
        f = open(log_file_path, 'a+', encoding='utf-8')
        for row in issues.issue_list:
            original_row = 'ORI |'
            for x, item in enumerate(row.original_row): 
                if x in row.mismatched_columns_indexes:
                    original_row += f' <|{item}|>'
                else:
                    original_row += f' {item}'
            uploaded_row = 'UPL |'
            for x, item in enumerate(row.uploaded_row): 
                if x in row.mismatched_columns_indexes:
                    uploaded_row += f' <|{item}|>'
                else:
                    uploaded_row += f' {item}'
            f.write(f'{original_row}\n')
            f.write(f'{uploaded_row}\n')
            f.write(f'\n')
        f.close()
        return

    # Writing logic for excel
    def write_to_excel():
        if not os.path.isfile(log_file_path):
            wb = Workbook()
            wb.save(log_file_path)
        sheet = wb.active # Create a pointer to the current sheet

        # Insert the fields based on the original file
        for column_no, field in enumerate(issues.original_fields):
            sheet.cell(row=1, column=column_no+2).value = field

        # Insert the issues 
        current_row = 2
        for issue in issues.issue_list:
            current_column = 1

            # Place the identifier in the first column 
            sheet.cell(row=current_row, column=current_column).value = 'ORI'
            
            # For every following column, put in the values from the original sheet
            current_column = 2
            for value in issue.original_row: 
                sheet.cell(row=current_row, column=current_column).value = value
                current_column += 1 

            # Highlight the columns that have discrepancies
            for mismatched_column_index in issue.mismatched_columns_indexes:
                # Ensure to offset the column as the first column is always the identifier
                sheet.cell(row=current_row, column=mismatched_column_index+2).fill = styles.PatternFill(fill_type="solid", fgColor="26B688")
            
            # Place identifier for the second column 
            current_column = 1
            current_row += 1
            sheet.cell(row=current_row, column=current_column).value = 'UPL'

            # For every following column, put values from the uploaded sheet
            current_column = 2
            for value in issue.uploaded_row: 
                sheet.cell(row=current_row, column=current_column).value = value
                current_column += 1    
            
            # Highlight the columns that have discrepancies
            for mismatched_column_index in issue.mismatched_columns_indexes:
                # Ensure to offset the column as the first column is always the identifier
                sheet.cell(row=current_row, column=mismatched_column_index+2).fill = styles.PatternFill(fill_type="solid", fgColor="FF8080")

            current_row += 1

        wb.save(log_file_path)
        return

    # Exit prematurely if there's no issues to write
    if len(issues.issue_list) == 0: 
        return

    # Get the exact path and file to log 
    log_file_path = f'{output_dir}/issues_{issues.name}_{time.strftime("%Y_%m_%d_%H_%M_%S", time.gmtime())}'
    if use_excel: 
        log_file_path += '.xlsx'
        write_to_excel()
    else:
        log_file_path += '.txt'
        write_to_text()
    
    if progress_status:
        progress_status(f'Wrote issues to {log_file_path}')

    return

def write_multiple_issues(issue_main_list: list[ISSUES_MAIN], progress_bar = None, progress_status = None, output_dir: str = "", output_to_excel: bool = True) -> None:
    folder_name = str(f'issues_{time.strftime("%Y_%m_%d_%H_%M_%S", time.gmtime())}')

    if progress_status != None: 
        progress_status(f'Logging issues to {folder_name}')
    if progress_bar: 
        progress_bar(0)

    # Check if there are any valid issues to log
    has_issues_in_general = any(item.has_issues() for item in issue_main_list)
    # If no valid issues to log, exit the function and tell user 
    if progress_status != None and has_issues_in_general == False: 
        progress_status(f'Done! No issues found ᕙ(⇀‸↼‶)ᕗ')
        return
    
    # Create a folder for the current log job
    folder_path_for_job = f'{output_dir}/{folder_name}'
    if not os.path.isdir(folder_path_for_job):
        os.mkdir(folder_path_for_job)

    # Creating a queue of tasks to be completed 
    task_queue = Queue()

    # Create a lock for thread-safe progress updates
    p = Progress(len(issue_main_list), progress_bar)

    # Put each issue into the queue 
    for issue in issue_main_list:
        task_queue.put(issue)

    def worker(task_queue: Queue, p:Progress):
        """Worker thread function that processes tasks from the queue."""
        while True:
            try:
                issue = task_queue.get(False)  # Try to get an item without blocking
                write_issues(issue, output_dir=folder_path_for_job, use_excel=output_to_excel)
                p.update_progress()
                task_queue.task_done()
            except Empty: 
                break

    # Create and start threads
    num_threads = min(len(issue_main_list), os.cpu_count())  # Use maximum available cores
    threads = [threading.Thread(target=worker, args=(task_queue,p)) for _ in range(num_threads)]
    for thread in threads:
        thread.start()

    # Wait for all threads to finish
    for thread in threads:
        thread.join()

    if progress_status != None: 
        progress_status(f'Done! Check issues_{folder_name}.')

    return

def write_multiple_issues_single_threaded(issue_main_list: list[ISSUES_MAIN], progress_bar = None, progress_status = None, output_dir: str = "", output_to_excel: bool = True) -> None:
    has_issues_in_general = any(item.has_issues() for item in issue_main_list)
    if has_issues_in_general:
        folder_path_for_job = str(f'{output_dir}/issues_{time.strftime("%Y_%m_%d_%H_%M_%S", time.gmtime())}')
        if not os.path.isdir(folder_path_for_job):
            os.mkdir(folder_path_for_job)
        for i, issue in enumerate(issue_main_list): 
            if issue.has_issues():
                if progress_bar != None and progress_status != None:
                    progress_bar(i/len(issue_main_list)*100)
                    progress_status(f'Creating issue log for {issue.name}')
                write_issues(issue, output_dir=folder_path_for_job, use_excel=output_to_excel)
        if progress_status != None: 
            progress_status(f'Done! Check {folder_path_for_job}.')
    if progress_status != None and has_issues_in_general == False: 
        progress_status(f'Done! No issues found ᕙ(⇀‸↼‶)ᕗ')
    return 

def compare_csv_folders(uploaded_folder_path: str, 
                       original_folder_path: str, 
                       progress_to_show_in_gui = None,
                       status_to_show_in_gui = None,
                       uploaded_file_identifying_field_index: int = 0,
                       original_file_identifying_field_index: int = 0,
                       ignore_leading_and_trailing_whitespaces: bool = False) -> list[ISSUES_MAIN]:
    
    if status_to_show_in_gui: 
        status_to_show_in_gui('Processing files.')
    
    # Find all valid sheets
    uploaded_file_names = [item for item in os.listdir(uploaded_folder_path) if item.endswith('.csv') or item.endswith('.xlsx')]   
    original_file_names = [item for item in os.listdir(original_folder_path) if item.endswith('.csv') or item.endswith('.xlsx')]
    # Hash the identifier to the original sheet file path
    original_file_paths_hash = {} 
    for item in original_file_names:
        original_file_paths_hash[item.split('-')[0]] =  f'{original_folder_path}/{item}'

    q = Queue()
    p = Progress(len(uploaded_file_names), progress_to_show_in_gui)
    issue_list_append_lock = threading.Lock()
    issues_list: list[ISSUES_MAIN] = []
    def add_to_issues_list(issues: ISSUES_MAIN):
        with issue_list_append_lock:
            issues_list.append(issues)
        return
    
    for uploaded_file_name in uploaded_file_names:
        q.put(uploaded_file_name)
    
    def worker(task_queue: Queue, p:Progress):  
        while True:
            try: 
                uploaded_file_name = task_queue.get(False) 
                uploaded_file_path = f'{uploaded_folder_path}/{uploaded_file_name}'
                try:
                    original_file_path = f'{original_file_paths_hash[uploaded_file_name.split("-")[0]]}'
                except KeyError: 
                    raise KeyError(f'STOP: "{uploaded_file_name}", cannot find any original file that starts with "{ori_file_name.split("-")[0]}".')
                result = find_discrepancies(uploaded_file_path=uploaded_file_path,
                            original_file_path=original_file_path,
                            uploaded_file_identifying_field_index=uploaded_file_identifying_field_index,
                            original_file_identifying_field_index=original_file_identifying_field_index,
                            ignore_leading_and_trailing_whitespaces=ignore_leading_and_trailing_whitespaces)
                add_to_issues_list(result)
                p.update_progress() 
                task_queue.task_done()
            except Empty: 
                break
    
    # Create and start threads
    num_threads = min(len(uploaded_file_names), os.cpu_count())  # Use maximum available cores
    threads = [threading.Thread(target=worker, args=(q,p)) for _ in range(num_threads)]
    for thread in threads:
        thread.start()

    # Wait for all threads to finish
    for thread in threads:
        thread.join()
    
    if status_to_show_in_gui: 
        status_to_show_in_gui('Finished processing files.')

    return issues_list

def compare_csv_folders_single_threaded(uploaded_folder_path: str, 
                       original_folder_path: str, 
                       progress_to_show_in_gui = None,
                       status_to_show_in_gui = None,
                       uploaded_file_identifying_field_index: int = 0,
                       original_file_identifying_field_index: int = 0,
                       ignore_leading_and_trailing_whitespaces: bool = False) -> list[ISSUES_MAIN]:
    
    uploaded_file_names = [item for item in os.listdir(uploaded_folder_path) if item.endswith('.csv') or item.endswith('.xlsx')]   
    original_file_names = [item for item in os.listdir(original_folder_path) if item.endswith('.csv') or item.endswith('.xlsx')]
    original_file_paths_hash = {} 
    
    for item in original_file_names:
        original_file_paths_hash[item.split('-')[0]] =  f'{original_folder_path}/{item}'

    issues_list: list[ISSUES_MAIN] = []
    
    for i, uploaded_file_name in enumerate(uploaded_file_names):
        uploaded_file_path = f'{uploaded_folder_path}/{uploaded_file_name}'
        try:
            original_file_path = f'{original_file_paths_hash[uploaded_file_name.split("-")[0]]}'
        except KeyError: 
            raise KeyError(f'STOP: "{uploaded_file_name}", cannot find any original file that starts with "{ori_file_name.split("-")[0]}".')

        issues_list.append(find_discrepancies(uploaded_file_path=uploaded_file_path,
                           original_file_path=original_file_path,
                           status_to_show_in_gui=status_to_show_in_gui,
                           progress_to_show_in_gui=progress_to_show_in_gui,
                           uploaded_file_identifying_field_index=uploaded_file_identifying_field_index,
                           original_file_identifying_field_index=original_file_identifying_field_index,
                           ignore_leading_and_trailing_whitespaces=ignore_leading_and_trailing_whitespaces))
        status_to_show_in_gui(f'Completed processing of {uploaded_file_name}')

    return issues_list

if __name__ == "__main__": 
    def demo():
        try: 
            uploaded_file_identifying_field_index = int(sys.argv[3])
            original_file_identifying_field_index = int(sys.argv[4])
        except:
            uploaded_file_identifying_field_index = 0
            original_file_identifying_field_index = 0

        # Give error and quit script if files cannot be found
        if (os.path.isfile(ori_file_name) and os.path.isfile(uploaded_file_name)):
            issues = find_discrepancies(uploaded_file_name,
                                        ori_file_name, 
                                        uploaded_file_identifying_field_index=uploaded_file_identifying_field_index,
                                        original_file_identifying_field_index=original_file_identifying_field_index)
            write_issues(issues, 'test_data')
            return

        if (os.path.isdir(ori_file_name) and os.path.isdir(uploaded_file_name)):
            res = compare_csv_folders(uploaded_folder_path=uploaded_file_name,
                                            original_folder_path=ori_file_name,
                                            uploaded_file_identifying_field_index=uploaded_file_identifying_field_index,
                                            original_file_identifying_field_index=original_file_identifying_field_index)
            write_multiple_issues(issue_main_list=res, output_dir='test_data')
            return

        raise Exception(f'Files {ori_file_name} or {uploaded_file_name} cannot be found. Terminating script.')

    # Run without GUI or use in notebook 
    name_of_script = sys.argv[0]
    ori_file_name = sys.argv[1]
    uploaded_file_name = sys.argv[2]
    demo()
