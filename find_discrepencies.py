"""
Takes in two line arguments:
1 - original csv file to be compared against
2 - seconday csv file which should match the first ASIDE from order
Compares both for discrepencies. If any are found they are noted into a file 'issues.txt' with their location in Excel format. 

Assumptions:
- All fields should be the same
- The unique identifier of each row is the first column 
"""
import csv 
import sys
import os.path
# from tqdm import tqdm
import time
from enum import Enum
from openpyxl import Workbook, styles, load_workbook

class NATURE_OF_ISSUES(Enum): 
    OK = "No issues found."
    DISCREPENCY = "Discrepencies found, please check the issue log."
    FIELDS_MISMATCH = "Columns fields don't match, please check the issue log."
    FIELDS_LENGTH_MISMATCH = "The number of columns in each csv file don't match. Please check for any trailing commas with notepad."
    
class ISSUE_ITEM:
    def __init__(self, original_row: list, uploaded_row: list, mismatched_columns_indexes: list, mismatched_columns_indexes_upl: list) -> None:
        self.original_row = original_row
        self.uploaded_row = uploaded_row
        self.mismatched_columns_indexes = mismatched_columns_indexes
        self.mismatched_columns_indexes_upl = mismatched_columns_indexes_upl
    
class ISSUES_MAIN:
    def __init__(self) -> None:
        self.nature_of_issues: list[NATURE_OF_ISSUES] = [] 
        self.name = ""
        self.issue_list: list[ISSUE_ITEM] = []
        self.original_fields: list[str] = []
        self.uploaded_fields: list[str] =[] 
        self.uploaded_hashed_fields_idxs: dict = {} #original field: uploaded field index 

    def has_issues(self) -> bool:
        return len(self.nature_of_issues) > 0 
    
    def update_uploaded_hashed_fields_idxs(self) -> dict:
        for ori_field in self.original_fields:
            for upl_field_idx, upl_field in enumerate(self.uploaded_fields):
                if upl_field == ori_field:
                    self.uploaded_hashed_fields_idxs[ori_field] = upl_field_idx
        return self.uploaded_hashed_fields_idxs
    
    def insert_issue_missing_uploaded_row(self, original_row: list, value_of_identifier: str, column_of_identifier: int = 0):
        row_to_indicate_missing_row = []
        for i in range(len(original_row)):
            if i != column_of_identifier:
                row_to_indicate_missing_row.append('MISSING')
            else: 
                row_to_indicate_missing_row.append(value_of_identifier)
        issue_item = ISSUE_ITEM(original_row, row_to_indicate_missing_row, [column_of_identifier], [column_of_identifier])
        self.issue_list.append(issue_item)
        return issue_item

    def insert_issue(self, original_row: list, uploaded_row: list, columns_where_discrepency_is_found: list[int]):
        # TODO remove redundant columns_where_discrepency_is_found_upl
        # columns_where_discrepency_is_found_upl = []
        # for column_idx in columns_where_discrepency_is_found:
        #     columns_where_discrepency_is_found_upl.append(self.uploaded_hashed_fields_idxs[self.original_fields[column_idx]])
        # Uploaded row mapped to original row's columns 
        mapped_uploaded_row = [] 
        for ori_field in self.original_fields:
            upl_field_idx = self.uploaded_hashed_fields_idxs[ori_field]
            value_from_upl_sheet = uploaded_row[upl_field_idx]
            mapped_uploaded_row.append(value_from_upl_sheet)
        issue_item = ISSUE_ITEM(original_row, mapped_uploaded_row, columns_where_discrepency_is_found, columns_where_discrepency_is_found)
        self.issue_list.append(issue_item)
        return issue_item

    def insert_missing_column(self, original_columns, uploaded_columns, columns_missing_from_uploaded: list[str]):
        issue_item = ISSUE_ITEM(original_row=original_columns, 
                                 uploaded_row=uploaded_columns, 
                                 mismatched_columns_indexes=columns_missing_from_uploaded, 
                                 mismatched_columns_indexes_upl=columns_missing_from_uploaded)
        self.issue_list.append(issue_item)
        return issue_item

def xlsx_to_csv(excel_file_path: str) -> str:
    wb = load_workbook(excel_file_path)
    sh = wb.active.values
    fields = [] 
    rest = []
    for i, row in enumerate(sh): 
        if i == 0: 
            fields = row 
        else: 
            rest.append(row)
    wb.close()
    return fields, rest 

def find_discrepencies(uploaded_file_path: str, 
                       original_file_path: str, 
                       progress_to_show_in_gui = None,
                       status_to_show_in_gui = None,
                       uploaded_file_identifiying_field_index: int = 0,
                       original_file_identifiying_field_index: int = 0) -> ISSUES_MAIN:
    """
    Finds the difference between two CSV files. 

    Args: 
        uploaded_file_path: The path that points to the CSV file which should match the original. 
        original_file_path: The path that points to the CSV file which is meant to be compared against (or in other words, this CSV file is the source of truth).
        progress_to_show_in_gui: A function from the GUI Python file which accepts the current percentage progress of the script as an integer.

    Returns: 
        A list containing strings of the discrepencies found.
    """
    # Update status
    if status_to_show_in_gui:
        status_to_show_in_gui('Setting up...')

    # If Excel file given, convert to CSV first
    if uploaded_file_path.split('.')[-1] == 'csv': 
        # Read uploaded csv file
        f_uploaded = open(uploaded_file_path, 'r')
        uploaded_csv_reader = csv.reader(f_uploaded)

        # Field names 
        fields_uploaded_csv = next(uploaded_csv_reader)
    else:
        fields_uploaded_csv, uploaded_csv_reader = xlsx_to_csv(uploaded_file_path)
    if original_file_path.split('.')[-1] == 'csv': 
        # Read original csv file
        f_ori = open(original_file_path, 'r')
        ori_csv_reader = csv.reader(f_ori)
        
        # Field names 
        fields_ori_csv = next(ori_csv_reader)
    else:
        fields_ori_csv, ori_csv_reader = xlsx_to_csv(original_file_path)

    # Initialise issues custom class to hold all the found issues (if any)
    issues = ISSUES_MAIN()
    issues.name = original_file_path.split('/')[-1].split('.')[0]
    # Save the fields for later use 
    issues.original_fields = fields_ori_csv
    issues.uploaded_fields = fields_uploaded_csv

    # Checking if every field in the original csv exists in the uploaded csv - If failed will not proceed with rest of check, return issue log immediately
    mismatched_fields = []
    for i in range(len(fields_ori_csv)):
        if fields_ori_csv[i] not in fields_uploaded_csv:
            issues.nature_of_issues.append(NATURE_OF_ISSUES.FIELDS_MISMATCH)
            mismatched_fields.append(i)
    if NATURE_OF_ISSUES.FIELDS_MISMATCH in issues.nature_of_issues: 
        issues.insert_missing_column(fields_ori_csv, fields_uploaded_csv, mismatched_fields)
        # TODO Toggle for checking regardless of fields being mismatched
        return issues
    
    # Update status
    if status_to_show_in_gui:
        status_to_show_in_gui('Caching uploaded csv...')

    # Hash the uploaded csv file based on its key value
    uploaded_hashed_csv = {}
    for row in uploaded_csv_reader:
        uploaded_hashed_csv[row[uploaded_file_identifiying_field_index]] = row
    
    # Hash the indexs of the uploaded csv fields
    uploaded_hashed_fields_index = issues.update_uploaded_hashed_fields_idxs()

    # Close the uploaded csv file, it's no longer needed as its now been hashed into memory
    try:
        f_uploaded.close()
    except UnboundLocalError: 
        pass

    # Update status
    if status_to_show_in_gui:
        status_to_show_in_gui('Comparing csvs...')

    # For each row in the original csv file
    # Check if it exists in the uploaded csv file
    # If it exists, compare that row of values to the original 
    previous_update = 0 # For GUI 
    # pbar = tqdm(total=len(uploaded_hashed_csv), desc="Comparing rows")

    # Each ROW 
    for row_num, row_from_ori_csv in enumerate( ori_csv_reader ):
        if row_from_ori_csv[original_file_identifiying_field_index] not in uploaded_hashed_csv:
            issues.insert_issue_missing_uploaded_row(row_from_ori_csv, row_from_ori_csv[original_file_identifiying_field_index], original_file_identifiying_field_index)
            continue

        row_from_uploaded_csv = uploaded_hashed_csv[row_from_ori_csv[original_file_identifiying_field_index]]

        mismatched_fields = []
        # for col_num in range(len(row_from_ori_csv)):
        #     if row_from_uploaded_csv[col_num] != row_from_ori_csv[col_num]:
        #         mismatched_fields.append(col_num)

        # Each COLUMN (CELL)
        for col_num in range(len(fields_ori_csv)):
            cell_from_ori_csv = row_from_ori_csv[col_num]
            cell_from_upl_csv = row_from_uploaded_csv[uploaded_hashed_fields_index[fields_ori_csv[col_num]]]
            if cell_from_ori_csv != cell_from_upl_csv:
                mismatched_fields.append(col_num)
        if len(mismatched_fields) > 0: 
            issues.insert_issue(row_from_ori_csv,row_from_uploaded_csv,mismatched_fields)
        
        # GUI and TQDM progress bars
        # pbar.update(1)
        progress_in_percentage = min( int(row_num / len(uploaded_hashed_csv)*100) + 1, 100)
        if progress_to_show_in_gui != None and previous_update != progress_in_percentage:
            progress_to_show_in_gui(progress_in_percentage)
            previous_update = progress_in_percentage
    
    # Mark issues found 
    if len(issues.issue_list): issues.nature_of_issues.append(NATURE_OF_ISSUES.DISCREPENCY)  

    # Close the original csv file, we've read through everything 
    try:
        f_ori.close()
    except UnboundLocalError: 
        pass

    return issues

def write_issues(issues: ISSUES_MAIN, output_dir: str = "", use_excel: bool = False, progress_bar = None, progress_status = None) -> None: 

    # Writing logic for a text file
    def write_to_text():
        f = open(log_file_path, 'a+')
        for row in issues.issue_list:
            original_row = 'ORI |'
            for x, item in enumerate(row.original_row): 
                if x in row.mismatched_columns_indexes:
                    original_row += f' <|{item}|>'
                else:
                    original_row += f' {item}'
            uploaded_row = 'UPL |'
            for x, item in enumerate(row.uploaded_row): 
                if x in row.mismatched_columns_indexes_upl:
                    uploaded_row += f' <|{item}|>'
                else:
                    uploaded_row += f' {item}'
            f.write(f'{original_row}\n')
            f.write(f'{uploaded_row}\n')
            f.write(f'\n')
        f.close()
        return

    # Writing logic for excel
    def write_to_excel():
        if not os.path.isfile(log_file_path):
            wb = Workbook()
            wb.save(log_file_path)
        sheet = wb.active # Create a pointer to the current sheet

        # Insert the fields based on the original file
        for column_no, field in enumerate(issues.original_fields):
            sheet.cell(row=1, column=column_no+2).value = field

        # Insert the issues 
        current_row = 2
        for issue in issues.issue_list:
            current_column = 1

            # Place the identifier in the first column 
            sheet.cell(row=current_row, column=current_column).value = 'ORI'
            
            # For every following column, put in the values from the original sheet
            current_column = 2
            for value in issue.original_row: 
                sheet.cell(row=current_row, column=current_column).value = value
                current_column += 1 

            # Highlight the columns that have discrepencies
            for mismatched_column_index in issue.mismatched_columns_indexes:
                # Ensure to offset the column as the first column is always the identifier
                sheet.cell(row=current_row, column=mismatched_column_index+2).fill = styles.PatternFill(fill_type="solid", fgColor="AFEEEE")
            
            # Place identifier for the second column 
            current_column = 1
            current_row += 1
            sheet.cell(row=current_row, column=current_column).value = 'UPL'

            # For every following column, put values from the uploaded sheet
            current_column = 2
            for value in issue.uploaded_row: 
                sheet.cell(row=current_row, column=current_column).value = value
                current_column += 1    
            
            # Highlight the columns that have discrepencies
            for mismatched_column_index in issue.mismatched_columns_indexes_upl:
                # Ensure to offset the column as the first column is always the identifier
                sheet.cell(row=current_row, column=mismatched_column_index+2).fill = styles.PatternFill(fill_type="solid", fgColor="AFEEEE")

            current_row += 1

        wb.save(log_file_path)
        return

    # Exit prematurely if there's no issues to write
    if len(issues.issue_list) == 0: 
        return

    # Get the exact path and file to log 
    log_file_path = f'{output_dir}/issues_{issues.name}_{time.strftime("%Y_%m_%d_%H_%M_%S", time.gmtime())}'
    if use_excel: 
        log_file_path += '.xlsx'
        write_to_excel()
    else:
        log_file_path += '.txt'
        write_to_text()
    
    return

def write_multiple_issues(issue_main_list: list[ISSUES_MAIN], progress_bar = None, progress_status = None, output_dir: str = "", output_to_excel: bool = True) -> None:
    folder_path_for_job = f'{output_dir}/issues_{time.strftime("%Y_%m_%d_%H_%M_%S", time.gmtime())}'
    if not os.path.isdir(folder_path_for_job):
        os.mkdir(folder_path_for_job)
    for issue in issue_main_list: 
        write_issues(issue, output_dir=folder_path_for_job, use_excel=output_to_excel)
    return 

def load_mapping_uploaded_to_original():
    f = open('mapping.csv', 'r')
    mapping_csv_reader = csv.reader(f)
    fields = next(mapping_csv_reader)
    mapping = {}
    for _, row in enumerate(mapping_csv_reader):
        original, uploaded = row[0], row[1]
        mapping[original] = uploaded
        mapping[uploaded] = original
    return mapping

def compare_csv_folders(uploaded_folder_path: str, 
                       original_folder_path: str, 
                       progress_to_show_in_gui = None,
                       status_to_show_in_gui = None,
                       uploaded_file_identifiying_field_index: int = 0,
                       original_file_identifiying_field_index: int = 0) -> list[ISSUES_MAIN]:
    
    uploaded_file_names = [item for item in os.listdir(uploaded_folder_path) if item.endswith('.csv') or item.endswith('.xlsx')]   
    original_file_names = [item for item in os.listdir(original_folder_path) if item.endswith('.csv') or item.endswith('.xlsx')]
    original_file_paths_hash = {} 
    
    for item in original_file_names:
        original_file_paths_hash[item.split('-')[0]] =  f'{original_folder_path}/{item}'

    issues_list: list[ISSUES_MAIN] = []
    
    for i, uploaded_file_name in enumerate(uploaded_file_names):
        uploaded_file_path = f'{uploaded_folder_path}/{uploaded_file_name}'
        try:
            original_file_path = f'{original_file_paths_hash[uploaded_file_name.split("-")[0]]}'
        except: 
            status_to_show_in_gui(f'STOP: "{uploaded_file_name}", cannot any original file that starts with "{uploaded_file_name.split("-")[0]}".')
            break

        status_to_show_in_gui(f'Processing file {uploaded_file_name}')
        issues_list.append(find_discrepencies(uploaded_file_path=uploaded_file_path,
                           original_file_path=original_file_path,
                           uploaded_file_identifiying_field_index=uploaded_file_identifiying_field_index,
                           original_file_identifiying_field_index=original_file_identifiying_field_index))
        status_to_show_in_gui(f'Completed processing of {uploaded_file_name}')

    return issues_list

if __name__ == "__main__": 
    # Get input parameters, if not use default names and column index for identifier
    try:
        name_of_script = sys.argv[0]
        ori_file_name = sys.argv[1]
        uploaded_file_name = sys.argv[2]
    except: 
        print(f'Using default filenames ori.csv and uploaded.csv with column index 0 as identifier\n\n')
        ori_file_name = 'ori.csv'
        uploaded_file_name = 'uploaded.csv'

    file = 'ori/SPTE-Pre-Upload Customer Master Data Trial Conversion - SPTE.xlsx'
    # try: 
    #     uploaded_file_identifiying_field_index = int(sys.argv[3])
    #     original_file_identifiying_field_index = int(sys.argv[4])
    # except:
    #     uploaded_file_identifiying_field_index = 0
    #     original_file_identifiying_field_index = 0

    # # Give error and quit script if files cannot be found
    # if not (os.path.isfile(ori_file_name) and os.path.isfile(uploaded_file_name)):
    #     raise Exception(f'Files {ori_file_name} or {uploaded_file_name} cannot be found. Terminating script.')
    
    # issues = find_discrepencies(uploaded_file_name,
    #                              ori_file_name, 
    #                              uploaded_file_identifiying_field_index=uploaded_file_identifiying_field_index,
    #                              original_file_identifiying_field_index=original_file_identifiying_field_index)
    # write_issues(issues.issue_list)
